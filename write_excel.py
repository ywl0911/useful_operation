# coding:utf8


from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be CONVERTED
import datetime

ws['A2'] = datetime.datetime.now()

# Save the file
wb.save('c:Users/ywl/Desktop/demo.xlsx')

print   1
